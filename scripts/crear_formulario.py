import os
from copy import copy

from openpyxl import Workbook, load_workbook

SOURCE_FILE = "Kit_Valor_Oculto_Master_FINAL.xlsx"
OUTPUT_FILE = "Kit_Valor_Oculto_Master_v2.xlsx"

SHEETS = [
    "s_Info",
    "Defaults_por_respuesta",
    "Riesgo_Total",
    "Calculos_Escenarios",
    "Reporte_1Pager_Data",
]


def copy_sheet(src_ws, dest_ws):
    """Copy cell values and styles from src_ws into dest_ws."""
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)


def apply_formulas(wb):
    """Insert formulas that depend on the selected response row."""
    # Defaults_por_respuesta
    dws = wb["Defaults_por_respuesta"]
    dws["B3"] = '=INDEX(Respuestas!A:AE, s_Info!$B$3, MATCH("tipo_ciudad (A/B/C/D)", Respuestas!$1:$1, 0))'
    dws["B4"] = '=INDEX(Respuestas!A:AE, s_Info!$B$3, MATCH("arquetipo (CHT/CBD/FRA/CMS/CLI/PRN)", Respuestas!$1:$1, 0))'
    dws["B5"] = '=INDEX(Respuestas!A:AE, s_Info!$B$3, MATCH("valor_mercado_hoy_MXN", Respuestas!$1:$1, 0))'

    for row in range(8, dws.max_row + 1):
        key = dws[f"A{row}"].value
        if key:
            dws[f"F{row}"] = f'=INDEX(Respuestas!A:AE, s_Info!$B$3, MATCH(A{row}, Respuestas!$1:$1, 0))'

    # Riesgo_Total
    rws = wb["Riesgo_Total"]
    rws["A4"] = "=INDEX(Respuestas!W:AE, s_Info!$B$3, 0)"

    # Calculos_Escenarios
    cws = wb["Calculos_Escenarios"]
    cws["B3"] = "=Defaults_por_respuesta!B5"
    cws["B4"] = "=Defaults_por_respuesta!G14"
    cws["B5"] = "=Defaults_por_respuesta!G8"

    # Reporte_1Pager_Data
    pws = wb["Reporte_1Pager_Data"]
    pws["B4"] = "=Calculos_Escenarios!B3"
    pws["B5"] = "=Calculos_Escenarios!B4"
    pws["D4"] = "=Riesgo_Total!A4"


def main():
    src = load_workbook(SOURCE_FILE)
    wb = Workbook()
    wb.remove(wb.active)

    for name in SHEETS:
        src_ws = src[name]
        dest_ws = wb.create_sheet(title=name)
        copy_sheet(src_ws, dest_ws)

    apply_formulas(wb)
    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    main()
